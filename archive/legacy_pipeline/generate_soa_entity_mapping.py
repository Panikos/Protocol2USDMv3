import openpyxl
import json
from collections import defaultdict

# --- CONFIG ---
EXCEL_PATH = "useful_material/USDM_CT.xlsx"
OUTPUT_PATH = "soa_entity_mapping.json"
SCHEMA_PATH = "USDM OpenAPI schema/USDM_API.json"
SHEET_ENTITIES = "DDF Entities&Attributes"
SHEET_VALUES = "DDF valid value sets"

# --- Load schema ---
print(f"[INFO] Loading schema from {SCHEMA_PATH}...")
with open(SCHEMA_PATH, 'r', encoding='utf-8') as f:
    schema_data = json.load(f)
schema_definitions = schema_data['components']['schemas']
print("[SUCCESS] Schema loaded.")

# --- Load workbook and sheets ---
print(f"[INFO] Loading Excel data from {EXCEL_PATH}...")
wb = openpyxl.load_workbook(EXCEL_PATH)
ws_entities = wb[SHEET_ENTITIES]
ws_values = wb[SHEET_VALUES]
print("[SUCCESS] Excel data loaded.")

# --- 1. Parse value sets from Excel for later use ---
value_sets = defaultdict(list)
header_row = 6
header_val = [cell.value for cell in ws_values[header_row]]
col_entity_val = header_val.index("Entity")
col_attr_val = header_val.index("Attribute")
col_codelist = header_val.index("Codelist C-code")
col_concept = header_val.index("Concept C-code")
col_term = header_val.index("Preferred Term (CDISC Submission Value)")
for row in ws_values.iter_rows(min_row=header_row+1):
    entity = row[col_entity_val].value
    attr = row[col_attr_val].value
    if entity and attr and row[col_term].value:
        value_sets[(entity, attr)].append({
            "term": row[col_term].value,
            "concept_c_code": row[col_concept].value,
            "codelist_c_code": row[col_codelist].value
        })

# --- 2. Create a lookup dictionary from Excel for enrichment data ---
excel_enrichment_data = defaultdict(dict)
header_ent = [cell.value for cell in ws_entities[1]]
col_entity_ent = header_ent.index("Entity Name")
col_role_ent = header_ent.index("Role")
col_ldm_ent = header_ent.index("Logical Data Model Name")
col_ccode_ent = header_ent.index("NCI C-code")
col_def_ent = header_ent.index("Definition")
for row in ws_entities.iter_rows(min_row=2):
    entity = row[col_entity_ent].value
    ldm_name = row[col_ldm_ent].value
    if not entity or not ldm_name or row[col_role_ent].value == "Entity":
        continue
    excel_enrichment_data[(entity, ldm_name)] = {
        "c_code": row[col_ccode_ent].value,
        "definition": row[col_def_ent].value,
        "role": row[col_role_ent].value
    }

# --- 3. Main Logic: Build mapping from the schema as the single source of truth ---
print("[INFO] Building entity mapping from schema...")
final_mapping = defaultdict(lambda: defaultdict(lambda: defaultdict(dict)))

for schema_name, schema_def in schema_definitions.items():
    if not schema_name.endswith("-Input"):
        continue
    entity_name = schema_name.replace("-Input", "")

    required_fields = schema_def.get('required', [])
    properties = schema_def.get('properties', {})

    for prop_name, prop_details in properties.items():
        enrichment = excel_enrichment_data.get((entity_name, prop_name), {})
        role = enrichment.get('role')
        if not role:
            is_ref = '$ref' in prop_details or ('items' in prop_details and '$ref' in prop_details['items'])
            role = "Relationship" if is_ref else "Attribute"

        field_info = {
            "name": prop_name,
            "c_code": enrichment.get('c_code'),
            "definition": prop_details.get('description', enrichment.get('definition', '')),
            "role": role,
            "required": prop_name in required_fields
        }

        allowed = value_sets.get((entity_name, prop_name), [])
        if allowed:
            field_info["allowed_values"] = allowed

        if role == "Attribute":
            final_mapping[entity_name]["attributes"][prop_name] = field_info
        elif role == "Relationship":
            final_mapping[entity_name]["relationships"][prop_name] = field_info
        elif role == "Complex Datatype Relationship":
            final_mapping[entity_name]["complex_datatype_relationships"][prop_name] = field_info

# --- 4. Inject special-case attributes for StudyDesign ---
# Per user feedback, trialType and observationalStudyModel are required for their use case.
study_design_attributes = {
    "trialType": {
        "name": "trialType", "role": "Attribute", "required": True,
        "definition": "A term that represents the nature of an interventional clinical trial..."
    },
    "observationalStudyModel": {
        "name": "observationalStudyModel", "role": "Attribute", "required": True,
        "definition": "A term that represents the nature of an observational study..."
    }
}
if "attributes" not in final_mapping['StudyDesign']:
    final_mapping['StudyDesign']['attributes'] = {}
final_mapping['StudyDesign']['attributes'].update(study_design_attributes)

# --- 4.5. Manually define SoA entities and relationships ---
# The standard schema processing doesn't capture the full semantic relationship
# for the Schedule of Activities. We inject it here to ensure the LLM
# understands the structure.
print("[INFO] Injecting manually defined SoA structure...")

# Define ActivityGroup as a conceptual entity for the LLM
final_mapping['ActivityGroup']['definition'] = "A conceptual grouping of activities that occur at the same timepoint (Encounter). In the protocol, this is often represented by a named row or section in the SoA table. In the USDM, this is represented by multiple 'ScheduledActivityInstance' objects sharing the same 'encounterId'."
final_mapping['ActivityGroup']['attributes'].update({
    "name": {"name": "name", "role": "Attribute", "required": False, "definition": "A descriptive name for the group of activities (e.g., 'Screening Procedures'). This may be inferred from the protocol's SoA table structure."},
    "id": {"name": "id", "role": "Attribute", "required": True, "definition": "A unique identifier for this conceptual group."}
})
final_mapping['ActivityGroup']['relationships'].update({
    "activities": {"name": "activities", "role": "Relationship", "required": True, "definition": "A list of 'ScheduledActivityInstance' objects that belong to this group."}
})

# Add descriptive definitions to key SoA entities processed from the schema
# Use 'PlannedTimepoint' as the canonical name for timepoints, as this is what the rest of the pipeline expects.
if 'Timing' in final_mapping:
    final_mapping['PlannedTimepoint'] = final_mapping.pop('Timing')
    final_mapping['PlannedTimepoint']['definition'] = "Represents a planned timepoint in the study schedule (e.g., 'Visit 1', 'Day 15'). This is the 'when' an activity occurs. In the SoA, this corresponds to a column header."

# Manually define ActivityTimepoint, which is missing from the OpenAPI spec but is the critical link between an activity and a timepoint.
final_mapping['ActivityTimepoint']['definition'] = "A link between an Activity and a PlannedTimepoint, signifying that the activity is performed at that time. In the SoA, this is represented by a checkmark or 'X' at the intersection of an activity row and a timepoint column."
final_mapping['ActivityTimepoint']['attributes'].update({
    "id": {"name": "id", "role": "Attribute", "required": True, "definition": "A unique identifier for this link."},
    "activityId": {"name": "activityId", "role": "Attribute", "required": True, "definition": "The ID of the Activity being performed."},
    "plannedTimepointId": {"name": "plannedTimepointId", "role": "Attribute", "required": True, "definition": "The ID of the PlannedTimepoint when the activity occurs."}
})
if 'Activity' in final_mapping:
    final_mapping['Activity']['definition'] = "A specific procedure or assessment performed during the study (e.g., 'Blood Draw', 'Questionnaire'). In the SoA, this corresponds to a row header."
if 'ScheduledActivityInstance' in final_mapping:
    final_mapping['ScheduledActivityInstance']['definition'] = "Represents the occurrence of a specific Activity at a specific Timepoint. This is the 'checkmark' in the SoA table, linking an Activity to a Timepoint via an Encounter."
if 'Encounter' in final_mapping:
    final_mapping['Encounter']['definition'] = "A crucial linking entity that associates one or more ScheduledActivityInstances (the 'what') with a single Timing object (the 'when'). It represents a single column in the SoA table."
if 'ScheduleTimeline' in final_mapping:
    final_mapping['ScheduleTimeline']['definition'] = "The main container for the entire Schedule of Activities. It holds all the timepoints (timings) and scheduled activities (instances)."

# --- 5. Manually define missing schema components ---
# The official USDM OpenAPI spec is missing definitions for Study, StudyTitle, and Code.
# We inject them here so the generated prompt correctly instructs the LLM to include them.
print("[INFO] Injecting manually defined schema components (Study, StudyTitle, Code)...")

# Define Code attributes
final_mapping['Code']['attributes'].update({
    "code": {"name": "code", "role": "Attribute", "required": True, "definition": "The coded value."},
    "decode": {"name": "decode", "role": "Attribute", "required": True, "definition": "The human-readable decode for the code."}
})

# Define StudyTitle attributes and relationships
final_mapping['StudyTitle']['attributes'].update({
    "text": {"name": "text", "role": "Attribute", "required": True, "definition": "The text of the study title."}
})
final_mapping['StudyTitle']['relationships'].update({
    "studyTitleType": {"name": "studyTitleType", "role": "Relationship", "required": True, "definition": "The type of the study title, as a Code object."}
})

# Define Study relationships
final_mapping['Study']['relationships'].update({
    "studyTitles": {"name": "studyTitles", "role": "Relationship", "required": True, "definition": "A list of titles for the study."}
})


# --- 6. Write to JSON ---
print(f"[INFO] Writing final mapping to {OUTPUT_PATH}...")
with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
    json.dump(final_mapping, f, indent=2, ensure_ascii=False)

print(f"[SUCCESS] Entity mapping written to {OUTPUT_PATH}")
